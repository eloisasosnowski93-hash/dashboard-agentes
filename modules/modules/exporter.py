"""Smart Product Finder v2.0 - Exportação CSV/Excel/JSON"""
import io, json, csv
from datetime import datetime
from typing import List

def export_to_csv(products: List[dict]) -> bytes:
    if not products: return b""
    output = io.StringIO()
    fields = ["nome","score","preco_custo","preco_sugerido","margem_pct","avaliacao","vendas",
              "dias_entrega","envia_brasil","selo_choice","categoria","decisao_ia","titulo_shopee","link"]
    w = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    w.writeheader()
    for p in products:
        cost = p.get("price",0); sug = p.get("ai_price_suggestion",0) or p.get("agent_price",0)
        margin = round(((sug-cost)/sug*100),1) if sug > 0 else 0
        w.writerow({"nome":p.get("name",""),"score":p.get("score",0),"preco_custo":f"R${cost:.2f}",
                    "preco_sugerido":f"R${sug:.2f}" if sug else "","margem_pct":f"{margin}%",
                    "avaliacao":p.get("rating",0),"vendas":p.get("sales",0),"dias_entrega":p.get("delivery_days",0),
                    "envia_brasil":"Sim" if p.get("ships_from_brazil") else "Não",
                    "selo_choice":"Sim" if p.get("choice_badge") else "Não","categoria":p.get("category",""),
                    "decisao_ia":p.get("ai_decision",""),"titulo_shopee":p.get("agent_title") or p.get("ai_shopee_title",""),
                    "link":p.get("link","")})
    return ("\ufeff" + output.getvalue()).encode("utf-8")

def export_to_excel(products: List[dict]) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Produtos"
        hf = PatternFill(start_color="FF6B35",end_color="FF6B35",fill_type="solid")
        hfont = Font(color="FFFFFF",bold=True,size=11)
        headers = ["Nome","Score","Custo","Venda","Margem%","Rating","Vendas","Prazo","BR","Choice","Cat","Decisão","Título Shopee","Link"]
        for col, h in enumerate(headers,1):
            cell = ws.cell(row=1,column=col,value=h); cell.fill=hf; cell.font=hfont
            cell.alignment=Alignment(horizontal="center")
        widths = [45,8,12,12,10,8,10,8,6,8,15,10,50,30]
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
        for rn, p in enumerate(products,2):
            cost=p.get("price",0); sug=p.get("agent_price") or p.get("ai_price_suggestion",0)
            margin=round(((sug-cost)/sug*100),1) if sug else 0; score=p.get("score",0) or 0
            row=[p.get("name",""),score,cost,sug or "",margin or "",p.get("rating",0),p.get("sales",0),
                 p.get("delivery_days",0),"✓" if p.get("ships_from_brazil") else "✗",
                 "✓" if p.get("choice_badge") else "✗",p.get("category",""),p.get("ai_decision",""),
                 p.get("agent_title") or p.get("ai_shopee_title",""),p.get("link","")]
            for col,val in enumerate(row,1):
                cell=ws.cell(row=rn,column=col,value=val)
                if col==2 and isinstance(val,(int,float)):
                    c="C8F7C5" if val>=75 else "FEF9E7" if val>=50 else "FDEBD0"
                    cell.fill=PatternFill(start_color=c,end_color=c,fill_type="solid")
        output=io.BytesIO(); wb.save(output); return output.getvalue()
    except ImportError:
        return export_to_csv(products)

def export_to_json(products: List[dict]) -> bytes:
    data = {"metadata":{"exported_at":datetime.now().isoformat(),"total":len(products),
                         "generator":"Smart Product Finder v2.0"},
            "products":[{"id":p.get("id"),"name":p.get("name"),"score":p.get("score",0),
                          "price":{"cost":p.get("price",0),"suggested":p.get("agent_price") or p.get("ai_price_suggestion",0)},
                          "metrics":{"rating":p.get("rating",0),"sales":p.get("sales",0)},
                          "ai_decision":p.get("ai_decision"),"link":p.get("link",""),
                          "agent_title":p.get("agent_title") or p.get("ai_shopee_title","")} for p in products]}
    return json.dumps(data,ensure_ascii=False,indent=2).encode("utf-8")

def get_export_filename(fmt, keyword="produtos"):
    ts=datetime.now().strftime("%Y%m%d_%H%M"); kw=keyword.replace(" ","_")[:20]
    ext={"csv":"csv","excel":"xlsx","json":"json"}.get(fmt,"txt")
    return f"smart_finder_{kw}_{ts}.{ext}"
