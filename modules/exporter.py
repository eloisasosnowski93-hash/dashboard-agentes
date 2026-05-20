"""
Smart Product Finder - Módulo de Exportação
Enzo (Gestor de Performance) usa estes exports para análise de ROI e planilhas de campanhas.
"""

import io
import json
import csv
from datetime import datetime
from typing import List


def export_to_csv(products: List[dict]) -> bytes:
    """
    Exporta produtos para CSV compatível com Excel brasileiro.
    Encoding UTF-8-BOM para abrir corretamente no Excel.
    """
    if not products:
        return b""

    output = io.StringIO()

    # Colunas para export (ordem otimizada para análise)
    fieldnames = [
        "nome", "score", "preco_custo", "preco_sugerido", "margem_pct",
        "avaliacao", "vendas", "dias_entrega", "envia_brasil", "selo_choice",
        "categoria", "decisao_ia", "titulo_shopee", "link", "coletado_em"
    ]

    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()

    for p in products:
        cost = p.get("price", 0)
        suggested = p.get("ai_price_suggestion", 0)
        margin = round(((suggested - cost) / suggested * 100), 1) if suggested > 0 else 0

        writer.writerow({
            "nome": p.get("name", ""),
            "score": p.get("score", 0),
            "preco_custo": f"R$ {cost:.2f}",
            "preco_sugerido": f"R$ {suggested:.2f}" if suggested else "",
            "margem_pct": f"{margin}%",
            "avaliacao": p.get("rating", 0),
            "vendas": p.get("sales", 0),
            "dias_entrega": p.get("delivery_days", 0),
            "envia_brasil": "Sim" if p.get("ships_from_brazil") else "Não",
            "selo_choice": "Sim" if p.get("choice_badge") else "Não",
            "categoria": p.get("category", ""),
            "decisao_ia": p.get("ai_decision", ""),
            "titulo_shopee": p.get("ai_shopee_title", ""),
            "link": p.get("link", ""),
            "coletado_em": p.get("collected_at", ""),
        })

    # UTF-8-BOM para compatibilidade com Excel brasileiro
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def export_to_excel(products: List[dict]) -> bytes:
    """
    Exporta para Excel (.xlsx) com formatação profissional.
    Usa openpyxl para criar planilha com cabeçalho colorido e larguras ajustadas.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Produtos Analisados"

        # Estilo do cabeçalho
        header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        center = Alignment(horizontal="center", vertical="center")

        headers = [
            "Nome do Produto", "Score", "Preço Custo",
            "Preço Sugerido", "Margem %", "Avaliação",
            "Vendas", "Prazo (dias)", "Brasil", "Choice",
            "Categoria", "Decisão IA", "Título Shopee", "Link"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # Larguras das colunas
        col_widths = [45, 8, 14, 14, 10, 10, 10, 12, 8, 8, 18, 12, 50, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[1].height = 22

        # Dados
        for row_num, p in enumerate(products, 2):
            cost = p.get("price", 0)
            suggested = p.get("ai_price_suggestion", 0)
            margin = round(((suggested - cost) / suggested * 100), 1) if suggested > 0 else 0
            score = p.get("score", 0) or 0

            row_data = [
                p.get("name", ""),
                score,
                cost,
                suggested or "",
                margin if margin else "",
                p.get("rating", 0),
                p.get("sales", 0),
                p.get("delivery_days", 0),
                "✓" if p.get("ships_from_brazil") else "✗",
                "✓" if p.get("choice_badge") else "✗",
                p.get("category", ""),
                p.get("ai_decision", ""),
                p.get("ai_shopee_title", ""),
                p.get("link", ""),
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)

                # Colorir score
                if col == 2 and isinstance(value, (int, float)):
                    if value >= 75:
                        cell.fill = PatternFill(start_color="C8F7C5", end_color="C8F7C5", fill_type="solid")
                    elif value >= 50:
                        cell.fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
                    elif value > 0:
                        cell.fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")

                # Colorir decisão
                if col == 12:
                    if value == "aprovado":
                        cell.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
                    elif value == "rejeitado":
                        cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

                # Linha alternada
                if row_num % 2 == 0 and col not in [2, 12]:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

        # Segunda aba: apenas aprovados
        approved = [p for p in products if p.get("ai_decision") == "aprovado" or p.get("is_approved")]
        if approved:
            ws2 = wb.create_sheet("✅ Aprovados")
            ws2.cell(row=1, column=1, value="Produtos Aprovados para Publicação")
            ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
            # Simplificado: só os essenciais
            for row_num, p in enumerate(approved, 3):
                ws2.cell(row=row_num, column=1, value=p.get("name", ""))
                ws2.cell(row=row_num, column=2, value=p.get("score", 0))
                ws2.cell(row=row_num, column=3, value=p.get("ai_shopee_title", ""))

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    except ImportError:
        # Fallback: retorna CSV se openpyxl não estiver instalado
        return export_to_csv(products)


def export_to_json(products: List[dict]) -> bytes:
    """
    Exporta para JSON estruturado com metadados.
    Útil para integração com outros sistemas ou APIs.
    """
    export_data = {
        "metadata": {
            "exported_at": datetime.now().isoformat(),
            "total_products": len(products),
            "total_approved": sum(1 for p in products if p.get("ai_decision") == "aprovado"),
            "generator": "Smart Product Finder v1.0"
        },
        "products": []
    }

    for p in products:
        export_data["products"].append({
            "id": p.get("id", ""),
            "name": p.get("name", ""),
            "score": p.get("score", 0),
            "price": {
                "cost": p.get("price", 0),
                "suggested": p.get("ai_price_suggestion", 0),
            },
            "metrics": {
                "rating": p.get("rating", 0),
                "sales": p.get("sales", 0),
                "delivery_days": p.get("delivery_days", 0),
                "ships_from_brazil": p.get("ships_from_brazil", False),
                "choice_badge": p.get("choice_badge", False),
            },
            "ai_analysis": {
                "decision": p.get("ai_decision", ""),
                "shopee_title": p.get("ai_shopee_title", ""),
            },
            "link": p.get("link", ""),
            "category": p.get("category", ""),
        })

    return json.dumps(export_data, ensure_ascii=False, indent=2).encode("utf-8")


def get_export_filename(format_type: str, keyword: str = "produtos") -> str:
    """Gera nome de arquivo com timestamp para download."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    safe_keyword = keyword.replace(" ", "_")[:20]
    extensions = {"csv": "csv", "excel": "xlsx", "json": "json"}
    ext = extensions.get(format_type, "txt")
    return f"smart_finder_{safe_keyword}_{timestamp}.{ext}"
